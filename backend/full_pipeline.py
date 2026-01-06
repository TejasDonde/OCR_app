import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from docling.document_converter import DocumentConverter
from openai import AzureOpenAI
from dotenv import load_dotenv


load_dotenv()

AZURE_API_KEY = os.getenv("AZURE_API_KEY")
AZURE_ENDPOINT = os.getenv("AZURE_ENDPOINT")
AZURE_API_VERSION = os.getenv("AZURE_API_VERSION")
AZURE_DEPLOYMENT = os.getenv("AZURE_DEPLOYMENT")

# Run: uvicorn main:app --reload --host 0.0.0.0 --port 8000
# ==========================================
# 1. CONFIGURATION & SETUP
# ==========================================

# Folder Structure
INPUT_FOLDER = "input_pdf"
TEMP_TXT_FOLDER = "temp_txt"
OUTPUT_JSON_FOLDER = "output_json"
OUTPUT_EXCEL_FOLDER = "output_excel"


# Azure OpenAI Credentials
# REPLACE THESE WITH YOUR ACTUAL CREDENTIALS
AZURE_ENDPOINT = AZURE_ENDPOINT
AZURE_API_KEY = AZURE_API_KEY
AZURE_API_VERSION = AZURE_API_VERSION
AZURE_DEPLOYMENT = "jss-gpt-4o"

# Embedded Schema (Reconstructed from your requirements to make script standalone)
# You can also load this from a file if you prefer, but this makes the script single-file.
JSON_SCHEMA = {
  "": "",
  "Company Code": "string",
  "Legal Entity Name": "string",
  "Vendor No": "string",
  "Vendor Name": "string",
  " ": "",
  "Subject": "string",
  "  ": "",
  "   ": [
    {
      "Invoice Date": "DD-MM-YYYY",
      "Invoice No": "string",
      "Purchase order No. if available": "string",
      "Invoice Amount": 0.0,
      "Invoice Currency": "string",
      "Invoice Due Date": "DD-MM-YYYY",
      "Remarks if any": "string"
    }
  ]
}

def setup_folders():
    """Ensure all necessary folders exist."""
    folders = [INPUT_FOLDER, TEMP_TXT_FOLDER, OUTPUT_JSON_FOLDER, OUTPUT_EXCEL_FOLDER]
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(folder)
            print(f"✓ Created folder: {folder}")

# ==========================================
# 2. MODULE: PDF TO TEXT (Docling)
# ==========================================

def convert_pdfs_to_text():
    """Converts PDFs from input folder to Markdown/Text."""
    print(f"\n{'='*60}")
    print("STEP 1: PDF → TEXT CONVERSION")
    print(f"{'='*60}")

    converter = DocumentConverter()
    files_processed = 0

    files = [f for f in os.listdir(INPUT_FOLDER) if f.lower().endswith(".pdf")]
    
    if not files:
        print("No PDF files found in input folder.")
        return

    for file_name in files:
        pdf_path = os.path.join(INPUT_FOLDER, file_name)
        base_name = os.path.splitext(file_name)[0]
        txt_path = os.path.join(TEMP_TXT_FOLDER, base_name + ".txt")

        print(f"Converting: {file_name}...")
        
        try:
            result = converter.convert(pdf_path)
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(result.document.export_to_markdown())
            files_processed += 1
        except Exception as e:
            print(f"✗ Error converting {file_name}: {e}")

    print(f"✓ converted {files_processed} file(s).")

# ==========================================
# 3. MODULE: TEXT TO JSON (Azure OpenAI)
# ==========================================

def extract_data_with_ai():
    """Extracts structured JSON from text files using Azure OpenAI."""
    print(f"\n{'='*60}")
    print("STEP 2: AI DATA EXTRACTION")
    print(f"{'='*60}")

    client = AzureOpenAI(
        azure_endpoint=AZURE_ENDPOINT,
        api_key=AZURE_API_KEY,
        api_version=AZURE_API_VERSION
    )

    # --- STRICT SYSTEM PROMPT (UNCHANGED) ---
    system_prompt = """You are a precise data extraction assistant specialized in parsing invoice statements from markdown/text format into structured JSON.


Key responsibilities:
- **CRITICAL: Do NOT calculate or sum any amounts. Extract numbers exactly as they appear in the source.**
- Extract invoice data with 100% accuracy
- Convert European number formats (1.234,56) to standard decimal format (1234.56)
- Normalize dates to ISO format (DD-MM-YYYY)
- Parse table structures to identify invoice rows
- Extract currency codes from amounts. If not given mention currency available anywhere in pdf, like in total amount.
- Identify and extract PO numbers when present. If "Purchase order No." is missing, use the value(number) from "Customer Ref", "REFERENCE", "Reference", "ref", or "PO" (case-insensitive).
- Preserve invoice order as they appear in source

- **Debit / Credit Handling (STRICT RULE):**
  • If a "Debit" column or debit value exists for an invoice row, ALWAYS use the Debit amount as "Invoice Amount"
  • COMPLETELY IGNORE any Credit values (Credit must never be used as Invoice Amount)
  • If Debit is missing or empty, use the document’s Total / Balance / Amount Due as "Invoice Amount"
  • Do NOT calculate, net, subtract, or combine Debit and Credit under any circumstance

- **Extract the Total row directly from the text - look for labels like "Total", "Total Balance", "Total Amount Due", "Amount Payable", "Total Balance Outstanding"**
- **CRITICAL: The Total row should have "Total" in the "Purchase order No. if available" field and empty strings "" in "Invoice Date" and "Invoice No" fields**
- Use exact field names: "Company Code", "Legal Entity Name", "Vendor No", "Vendor Name", "Subject"
- Use table headers: "Invoice Date", "Invoice No", "Purchase order No. if available", "Invoice Amount", "Invoice Currency", "Invoice Due Date", "Remarks if any"
- Include blank rows for spacing: "", " ", "  " before and around Subject
- Use "   " (three spaces) as key for table array - NO text label
- "Remarks if any" field is not mandatory to be filled add only if available keep blank otherwise.
- Return ONLY valid JSON with no additional text
- Follow the provided schema exactly
- Use empty strings "" for missing text fields and 0.0 for missing numeric amounts"""    
# Key responsibilities:
# - **CRITICAL: Do NOT calculate or sum any amounts. Extract numbers exactly as they appear in the source.**
# - Extract invoice data with 100% accuracy
# - Convert European number formats (1.234,56) to standard decimal format (1234.56)
# - Normalize dates to ISO format (DD-MM-YYYY)
# - Parse table structures to identify invoice rows
# - Extract currency codes from amounts. If not given mention currency available anywhere in pdf, like in total amount.
# - Identify and extract PO numbers when present. If "Purchase order No." is missing, use the value(number) from "Customer Ref", "REFERENCE", "Reference", "ref", or "PO" (case-insensitive).
# - Preserve invoice order as they appear in source
# - **Extract the Total row directly from the text - look for labels like "Total", "Total Balance", "Total Amount Due", "Amount Payable", "Total Balance Outstanding"**
# - **CRITICAL: The Total row should have "Total" in the "Purchase order No. if available" field and empty strings "" in "Invoice Date" and "Invoice No" fields**
# - Use exact field names: "Company Code", "Legal Entity Name", "Vendor No", "Vendor Name", "Subject"
# - Use table headers: "Invoice Date", "Invoice No", "Purchase order No. if available", "Invoice Amount", "Invoice Currency", "Invoice Due Date", "Remarks if any"
# - Include blank rows for spacing: "", " ", "  " before and around Subject
# - Use "   " (three spaces) as key for table array - NO text label
# - "Remarks if any" field is not mandatory to be filled add only if available keep blank otherwise.
# - Return ONLY valid JSON with no additional text
# - Follow the provided schema exactly
# - Use empty strings "" for missing text fields and 0.0 for missing numeric amounts"""

    files = [f for f in os.listdir(TEMP_TXT_FOLDER) if f.lower().endswith(".txt")]
    processed_count = 0
    error_count = 0

    for file_name in files:
        txt_path = os.path.join(TEMP_TXT_FOLDER, file_name)
        json_path = os.path.join(OUTPUT_JSON_FOLDER, os.path.splitext(file_name)[0] + ".json")
        
        print(f"Processing: {file_name}")

        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                markdown_content = f.read()

            # --- STRICT USER PROMPT (UNCHANGED) ---
            prompt = f"""You are a data-extraction assistant. I will give you a markdown/text representation of an invoice statement extracted from a PDF. Parse that text and return **only** a single JSON object matching the schema described below. Do not add extra fields, comments, or explanations — return raw JSON and nothing else.

Rules and mapping:
1. Top-level fields (strings):
   - "Company Code": company code (e.g. "CS1058"). If not present, set to "".
   - "Legal Entity Name": full legal entity name. If not present, set to "".
   - "Vendor No": vendor number (use as string). If not present, set to "".
   - "Vendor Name": vendor name. If not present, set to "".
   - "Subject": subject or title for the statement/document. If missing, set to "".

2. Table data: an array of invoice objects. Each invoice object must include these keys:
   - "Invoice Date": normalized to ISO format `"DD-MM-YYYY"` if a date is present. If unparseable, set to "".
   - "Invoice No": invoice/document identifier (string). If not present, set to "".
   - "Purchase order No. if available": PO number if explicitly present. If not found, also accept Number from fields labeled "Purchase order No.", "Customer Ref", "REFERENCE", "Reference", "REF", "ref", or "PO" (case-insensitive). If none are present, set to "".
   - "Invoice Amount": numeric value as a number (no thousands separators, decimal point `.`). Convert amounts like `1.200,00 EUR` or `7.144,65 EUR` to `1200.00` and `7144.65`. If no amount, set to `0.0`.
   - "Invoice Currency": currency code or currency text (e.g. "EUR"). If currency appears with the amount, extract it. If missing, set to currency symbol mention in Total (e.g. £ as GBP, $ as USD, € as EUR etc).
   - "Invoice Due Date": normalized to `"DD-MM-YYYY"` if present; else "".
   - "Remarks if any": any short free-text remarks found in the same row or details cell; else "".

3. Parsing guidance:
   - The input may contain columns presented as tables or inline lines. Detect rows that appear to be invoices.
   - When amounts have a comma as decimal separator and dots as thousands separators (European style), convert to a float with dot decimal.
   - If a row contains "PO", "Purchase order", "Customer Ref", "REFERENCE", "ref", or similar labels, capture the corresponding number as the Purchase order No. if available.
   - Treat "Customer Ref", "REFERENCE", "Reference", "ref", and "PO" exactly the same as Purchase order No. and only take Number not other text.
   - **CRITICAL: Extract the Total row EXACTLY as it appears in the text file. Look for labels like "Total", "Total Balance", "Total Amount Due", "Amount Payable", "Total Balance Outstanding", etc. DO NOT calculate or sum the total - extract it directly from the source.**
   - The Total row should be included as the LAST item in the table array with "Purchase order No. if available" set to "Total" and "Invoice Date" and "Invoice No" set to "".

4. Output format (exact keys, example types):
{{
  "": "",
  "Company Code": "string",
  "Legal Entity Name": "string",
  "Vendor No": "string",
  "Vendor Name": "string",
  " ": "",
  "Subject": "string",
  "  ": "",
  "   ": [
    {{
      "Invoice Date": "DD-MM-YYYY",
      "Invoice No": "string",
      "Purchase order No. if available": "string",
      "Invoice Amount": number,
      "Invoice Currency": "string",
      "Invoice Due Date": "DD-MM-YYYY",
      "Remarks if any": "string"
    }},
    {{
      "Invoice Date": "",
      "Invoice No": "",
      "Purchase order No. if available": "Total",
      "Invoice Amount": number,
      "Invoice Currency": "string",
      "Invoice Due Date": "",
      "Remarks if any": ""
    }}
  ]
}}

5. CRITICAL FORMATTING RULES:
   - Include empty string keys ("", " ", "  ", "   ") for spacing as shown
   - Leave one blank row at top ("")
   - Leave one blank row before Subject (" ")
   - Leave one blank row after Subject ("  ")
   - Table array key is "   " (three spaces) - NO label like "invoices"
   - The last row MUST be the Total row exactly as it appears in the txt file. DO NOT calculate or sum anything.
   - **IMPORTANT: For the Total row, put "Total" in "Purchase order No. if available" field and leave "Invoice Date" and "Invoice No" as empty strings ""**
   - **Use exact field names as specified**

6. Respond with **only the JSON object** and nothing else.

---

Target JSON Schema:
{json.dumps(JSON_SCHEMA, indent=2)}

---

Markdown Content to Extract:
{markdown_content}

---

Return ONLY the JSON object matching the schema above. No markdown, no code blocks, no explanations."""

            response = client.chat.completions.create(
                model=AZURE_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                response_format={"type": "json_object"}
            )

            result_content = response.choices[0].message.content.strip()
            
            # Clean markdown code blocks
            if result_content.startswith("```json"): result_content = result_content[7:]
            if result_content.startswith("```"): result_content = result_content[3:]
            if result_content.endswith("```"): result_content = result_content[:-3]
            result_content = result_content.strip()

            output_data = json.loads(result_content)

            # Reorder and validate structure
            ordered_data = {}
            ordered_data[""] = output_data.get("", "")
            ordered_data["Company Code"] = output_data.get("Company Code", "")
            ordered_data["Legal Entity Name"] = output_data.get("Legal Entity Name", "")
            ordered_data["Vendor No"] = output_data.get("Vendor No", "")
            ordered_data["Vendor Name"] = output_data.get("Vendor Name", "")
            ordered_data[" "] = output_data.get(" ", "")
            ordered_data["Subject"] = output_data.get("Subject", "")
            ordered_data["  "] = output_data.get("  ", "")

            # Handle Table Data and enforce Total Row logic
            table_data = output_data.get("   ", [])
            for invoice in table_data:
                if invoice.get("Invoice No", "").lower() == "total" or invoice.get("Invoice Date", "").lower() == "total":
                    invoice["Purchase order No. if available"] = "Total"
                    invoice["Invoice Date"] = ""
                    invoice["Invoice No"] = ""
            
            ordered_data["   "] = table_data

            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(ordered_data, f, indent=2, ensure_ascii=False)

            processed_count += 1
            print(f"✓ Extracted JSON for {file_name}")

        except Exception as e:
            print(f"✗ Error processing {file_name}: {e}")
            error_count += 1
    
    print(f"\nAI Extraction Complete: {processed_count} success, {error_count} errors.")

# ==========================================
# 4. MODULE: JSON TO EXCEL (Pandas/OpenPyXL)
# ==========================================

def convert_json_to_excel():
    """Convert generated JSON files to formatted Excel."""
    print(f"\n{'='*60}")
    print("STEP 3: EXCEL REPORT GENERATION")
    print(f"{'='*60}")

    json_files = [f for f in os.listdir(OUTPUT_JSON_FOLDER) if f.lower().endswith('.json')]
    
    if not json_files:
        print("No JSON files found to convert.")
        return

    success_count = 0
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for json_file in json_files:
        json_path = os.path.join(OUTPUT_JSON_FOLDER, json_file)
        excel_file = os.path.splitext(json_file)[0] + ".xlsx"
        excel_path = os.path.join(OUTPUT_EXCEL_FOLDER, excel_file)
        
        print(f"Generating Report: {excel_file}")

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            current_row = 1

            # Recursive function to write data
            def write_recursive(obj, indent=0):
                nonlocal current_row
                
                if isinstance(obj, dict):
                    for key, value in obj.items():
                        if key in ["$schema", "title", "type"]: continue

                        # TABLE LOGIC
                        if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                            df = pd.DataFrame(value)
                            
                            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                                # Detect Total Row
                                is_total_row = False
                                if r_idx > 0:
                                    for val in row_data:
                                        if str(val).strip().lower() == "total":
                                            is_total_row = True
                                            break
                                
                                # Insert 4 blank rows before Total
                                if is_total_row:
                                    num_columns = len(row_data)
                                    for _ in range(4):
                                        for c_idx in range(1, num_columns + 1):
                                            blank_cell = ws.cell(row=current_row, column=c_idx + 1, value="")
                                            blank_cell.border = thin_border
                                        current_row += 1
                                
                                # Write Row
                                for c_idx, cell_value in enumerate(row_data, 1):
                                    cell = ws.cell(row=current_row, column=c_idx + 1, value=cell_value)
                                    if r_idx == 0: cell.font = Font(bold=True)
                                    cell.border = thin_border
                                current_row += 1
                            current_row += 1

                        # DICT LOGIC
                        elif isinstance(value, dict):
                            title = ws.cell(row=current_row, column=2, value=key.upper())
                            title.font = Font(bold=True)
                            current_row += 1
                            write_recursive(value, indent + 1)

                        # LIST LOGIC
                        elif isinstance(value, list):
                            ws.cell(row=current_row, column=2, value=key)
                            ws.cell(row=current_row, column=3, value=str(value))
                            current_row += 1

                        # KEY-VALUE LOGIC
                        else:
                            ws.cell(row=current_row, column=2, value=key)
                            ws.cell(row=current_row, column=3, value=value)
                            current_row += 1

            write_recursive(data)

            # Column Auto-width
            for column in ws.columns:
                max_length = 0
                column_list = list(column)
                for cell in column_list:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_list[0].column_letter].width = min(max_length + 2, 50)

            wb.save(excel_path)
            success_count += 1
            
        except Exception as e:
            print(f"✗ Error converting {json_file}: {e}")

    print(f"✓ Excel Generation Complete: {success_count} files created.")

# ==========================================
# 5. MAIN EXECUTION
# ==========================================

if __name__ == "__main__":
    print("STARTING INVOICE PROCESSING PIPELINE...")
    
    # 1. Setup
    setup_folders()
    
    # 2. Check for Inputs
    if not os.path.exists(INPUT_FOLDER) or not os.listdir(INPUT_FOLDER):
        print(f"Please put your PDF files inside the '{INPUT_FOLDER}' directory and run this script again.")
    else:
        # 3. Execution Chain
        convert_pdfs_to_text()
        extract_data_with_ai()
        convert_json_to_excel()
        
        print(f"\n{'='*60}")
        print("PIPELINE FINISHED SUCCESSFULLY")
        print(f"Check the '{OUTPUT_EXCEL_FOLDER}' folder for your reports.")
        print(f"{'='*60}")

